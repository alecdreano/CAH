#############################################################################################################################################
#Author: Alexandre DREANO
#Last update: 2018-08-09
#Info:
#StillToDo: /
#############################################################################################################################################
import os
import json
import openpyxl

config = {
	"max_char": 200,
	"gap_sign": {
		"input": "_",
		"output": "_____",
	},
	"excel": {
		"sheet_names": {
			"completion": "CARTES_COMPLETION",
			"gapped": {
				"1": "CARTES_1_TROU",
				"2": "CARTES_2_TROUS",
				"3": "CARTES_3_TROUS",
			},
		},
		"columns_index": {
			"rate": "D",
			"content": "A",
			"tags": "C",
			"template": "E",
			"category": "B",
		},
	},
	"cards": {
		"completion": {
			"number": 200,
		},
		"gapped": {
			"1": {
				"number": 50,
			},
			"2": {
				"number": 30,
			},
			"3": {
				"number": 2,
			},
		},
	},
	"paths": {
		"original_cards": "DATA/ORIGINAL_CARDS/",
		"new_cards": "DATA/ORIGINAL_CARDS/",
		"output": "DATA/OUTPUT/FINAL_CARDS/data.json",
		"output_html": "DESIGN/assets/data/cards_content/data.js",
	},
};

### UTILS FUNCTIONS 
def getExcelFilename(path):
	filenames = [ filename for filename in os.listdir(path) if (filename.endswith('.xls') or filename.endswith('.xlsx'))];
	if len(filenames) == 0:
		raise ValueError("No Excel file found at the given path : " + str(path));
	elif len(filenames) > 1:
		raise ValueError("Several Excel files found at the given path : " + str(path));
	else:
		return filenames[0];

def openExcelFileFromPath(path):
	filepath = path + getExcelFilename(path);
	workbook = openpyxl.load_workbook(filepath, data_only=True);
	return workbook;

def Card(type, content, details, source):
	return {
		"content": content,
		"type": type,
		"source": source,
		"details": details,
	};

### GET ORIGINAL_CARDS
def Deck():
	return { 
		"completion" : [],
		"gapped": {
			"1": [],
			"2": [],
			"3": [],
		},
	};

### GET ORIGINAL_CARDS
original_cards = Deck();
original_wb = openExcelFileFromPath(config["paths"]["original_cards"]);
excel_iter_params = [ 
	{"sheet": config["excel"]["sheet_names"]["completion"], "type": "completion", "list": original_cards["completion"]},
	{"sheet": config["excel"]["sheet_names"]["gapped"]["1"], "type": "gapped", "list": original_cards["gapped"]["1"], "details": {"gaps": 1}},
	{"sheet": config["excel"]["sheet_names"]["gapped"]["2"], "type": "gapped", "list": original_cards["gapped"]["2"], "details": {"gaps": 2}},
	{"sheet": config["excel"]["sheet_names"]["gapped"]["3"], "type": "gapped", "list": original_cards["gapped"]["3"], "details": {"gaps": 3}},
];
for param in excel_iter_params:
	sheet = original_wb[param["sheet"]];
	for row in range(2, sheet.max_row + 1):
		content  = str(sheet[config["excel"]["columns_index"]["content"] + str(row)].value.encode("utf-8"));
		rate = int(sheet[config["excel"]["columns_index"]["rate"] + str(row)].value);
		tags = sheet[config["excel"]["columns_index"]["tags"] + str(row)].value;
		category = sheet[config["excel"]["columns_index"]["category"] + str(row)].value;
		template = sheet[config["excel"]["columns_index"]["template"] + str(row)].value;
		details = {
			"rate": rate,
			"tags": tags,
			"category": category,
			"template": template,
		};
		if "details" in param.keys():
			for key in param["details"].keys():
				details[key] = param["details"][key];
		card = Card(param["type"], content, details, "original");
		param["list"].append(card);

### NEW_CARDS
new_cards = Deck();
new_wb = openExcelFileFromPath(config["paths"]["new_cards"]);
excel_iter_params = [ 
	{"sheet": config["excel"]["sheet_names"]["completion"], "type": "completion", "list": new_cards["completion"]},
	{"sheet": config["excel"]["sheet_names"]["gapped"]["1"], "type": "gapped", "list": new_cards["gapped"]["1"], "details": {"gaps": 1}},
	{"sheet": config["excel"]["sheet_names"]["gapped"]["2"], "type": "gapped", "list": new_cards["gapped"]["2"], "details": {"gaps": 2}},
	{"sheet": config["excel"]["sheet_names"]["gapped"]["3"], "type": "gapped", "list": new_cards["gapped"]["3"], "details": {"gaps": 3}},
];
for param in excel_iter_params:
	sheet = original_wb[param["sheet"]];
	for row in range(2, sheet.max_row + 1):
		content  = str(sheet[config["excel"]["columns_index"]["content"] + str(row)].value.encode("utf-8"));
		template = sheet[config["excel"]["columns_index"]["template"] + str(row)].value;
		category = sheet[config["excel"]["columns_index"]["category"] + str(row)].value;
		details = {
			"tags": tags,
			"category": category,
			"template": template,

		};
		if "details" in param.keys():
			for key in param["details"].keys():
				details[key] = param["details"][key];
		card = Card(param["type"], content, details, "new");
		param["list"].append(card);

#SELECT CARDS FOR DECK
final_cards = Deck();
excel_iter_params = [ 
	{"limit": config["cards"]["completion"]["number"], "original_cards": original_cards["completion"], "new_cards": new_cards["completion"], "final_cards": final_cards["completion"]},
	{"limit": config["cards"]["gapped"]["1"]["number"], "original_cards": original_cards["gapped"]["1"], "new_cards": new_cards["gapped"]["1"], "final_cards": final_cards["gapped"]["1"]},
	{"limit": config["cards"]["gapped"]["2"]["number"], "original_cards": original_cards["gapped"]["2"], "new_cards": new_cards["gapped"]["2"], "final_cards": final_cards["gapped"]["2"]},
	{"limit": config["cards"]["gapped"]["3"]["number"], "original_cards": original_cards["gapped"]["3"], "new_cards": new_cards["gapped"]["3"], "final_cards": final_cards["gapped"]["3"]},
];
for param in excel_iter_params:
	count = 0;
	### TAKE ALL NEWS CARDS
	while count < param["limit"] and count < len(param["new_cards"]):
		param["final_cards"].append(param["new_cards"][count])
		count += 1;
	### FILL WITH ORIGINAL CARDS IF NOT ENOUGH CARDS
	if count < param["limit"]:
		# SORT BY RATING TO TAKE BEST CARDS
		sorted_original_cards = sorted(param["original_cards"], key=lambda card: card["details"]["rate"], reverse=True);
		count_0 = 0;
		while count < param["limit"] and count_0 < len(sorted_original_cards):
			param["final_cards"].append(sorted_original_cards[count_0])
			count += 1;
			count_0 += 1;

def extendMultipleLists(list_of_list):
	output = [];
	for list in list_of_list:
		output.extend(list);
	return output;

def getAllCards(deck):
	all_lists = [
		deck["completion"],
		deck["gapped"]["1"],
		deck["gapped"]["2"],
		deck["gapped"]["3"],
	];
	return extendMultipleLists(all_lists);

for card in getAllCards(final_cards):
	print(card)
	#GLOBAL CHECK 
	if len(card["content"]) > config["max_char"]:
			raise ValueError("Too many characters for the following card content : " + str(card["content"]));
	if card["details"]["template"] is None and card["details"]["category"] is not None:
		raise ValueError("Card without explicite template must have category : " + str(card["content"]));
	# gapped CARDS CHECK
	if card["type"] == "gapped":
		card["content"] = card["content"].replace(config["gap_sign"]["input"], config["gap_sign"]["output"]);
	# COMPLETION CARDS CHECK
	elif card["type"] == "completion":
		if card["content"].find(config["gap_sign"]["input"]) >= 0:
			raise ValueError("Completion card must no contains gap sign : " + str(card["content"]));
	else:
		raise ValueError("Unknown type of card : " + str(card["type"]));

with open(config["paths"]["output"], 'w') as outfile:
	outfile.write(str(json.dumps(final_cards, indent=4)));

with open(config["paths"]["output_html"], 'w') as outfile:
	js_code = 'const json_cards = ' + json.dumps(final_cards, indent=4);
	outfile.write(str(js_code));